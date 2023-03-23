import openpyxl
import pandas as pd
import os
from openpyxl.styles import PatternFill

df1 = pd.read_excel("/Users/ziweishi/Desktop/Book2.xlsx",sheet_name="Sheet1")
df2 = pd.read_excel("/Users/ziweishi/Desktop/Book2.xlsx",sheet_name="Sheet2")
fill_cell = PatternFill(patternType='solid',
                            fgColor='C64747')

s3_list = df1["s3_content"].to_list()
dy_list=df2["dy_content"].to_list()

print(len(s3_list))
print(len(dy_list))
wb = openpyxl.load_workbook("/Users/ziweishi/Desktop/Book2.xlsx")
ws2 = wb['Sheet2']
a=2
loss_s3=[]
exceed_s3=[]

for i in dy_list:
    if i not in s3_list:
        loc = "B"+str(a)
        print(loc)
        ws2[loc].fill=fill_cell


    a+=1

c=2
ws1 = wb['Sheet1']
for j in s3_list:
    if j not in dy_list:
        loc = "B"+str(c)
        print(loc)
        ws1[loc].fill=fill_cell

    c+=1

wb.save("/Users/ziweishi/Desktop/Book2.xlsx")